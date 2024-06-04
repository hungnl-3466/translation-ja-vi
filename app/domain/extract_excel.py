from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import json
from datetime import datetime

# Hàm chuyển đổi đối tượng datetime thành chuỗi
def datetime_to_string(value):
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    return value
# Đường dẫn tới file Excel
excel_file = "/media/benu/DATA/sun-asterisk/translation-ja-vi/xls_data/Merge_cell_[1356]_ImageProcessingPoC_機能一覧.xlsx"

# Load workbook
wb = load_workbook(excel_file)

# Khởi tạo một danh sách để lưu trữ dữ liệu
data = []

# Lấy ra tất cả các sheets trong workbook
sheets = wb.sheetnames

# Lặp qua từng sheet để trích xuất dữ liệu
for sheet_name in sheets:
    sheet = wb[sheet_name]
    
    # Khởi tạo một danh sách để lưu trữ dữ liệu từ mỗi dòng trong sheet
    sheet_data = []
    
    # Lặp qua từng dòng trong sheet
    for row in sheet.iter_rows(values_only=True):
        # Khởi tạo một từ điển để lưu trữ dữ liệu từ mỗi ô trong dòng
        row_data = {}
        
        # Lặp qua từng ô trong dòng và lưu trữ dữ liệu vào từ điển
        for cell, value in zip(sheet[1], row):
            if isinstance(value, datetime):
                # Nếu giá trị là datetime, chuyển đổi thành chuỗi
                value = value.strftime("%Y-%m-%d %H:%M:%S")
                
                
            if value is not None:
                column_letter = get_column_letter(cell.column)
                row_data[f"{sheet[cell.coordinate].value} ({column_letter})"] = value
        
        # Kiểm tra xem dòng có chứa ít nhất một giá trị khác None không trước khi thêm vào danh sách dữ liệu của sheet
        if any(value is not None for value in row_data.values()):
            sheet_data.append(row_data)
    
    # Thêm danh sách dữ liệu của sheet vào danh sách tổng thể nếu có ít nhất một dòng chứa giá trị khác None
    if sheet_data:
        sheet_data_dict = {sheet_name: sheet_data}
        with open('{}.json'.format(sheet_name), 'w', encoding='utf-8') as f:
            json.dump(sheet_data_dict, f,ensure_ascii=False, indent=4)
        data.append(sheet_data_dict)

# In dữ liệu dưới dạng dictionary
# print(data)
# for data_sheet in data:
#     # for key in data_sheet.keys():
#     sheet
#     # for data_dict in 
#         # print(data_sheet[key])
#     with open('dict_slide_text_featuure_ver_5.json', 'w', encoding='utf-8') as f:
#         json.dump(data_sheet, f,ensure_ascii=False, indent=4)
#     print("-------------------------------------------")
