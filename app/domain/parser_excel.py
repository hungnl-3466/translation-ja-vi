import pandas as pd
from openpyxl import load_workbook
data = {
    "アプリ機能一覧": [
        {
            "No (A)": "No",
            "カテゴリ\n（画面名） (B)": "カテゴリ\n（画面名）",
            "機能名 (C)": "機能名",
            "機能説明 (D)": "機能説明",
            "補足 (E)": "補足",
            "55555555 (F)": 55555555
        },
        {
            "No (A)": 1,
            "カテゴリ\n（画面名） (B)": "スプラッシュ",
            "機能名 (C)": "導入画面",
            "機能説明 (D)": "1秒間ロゴとアプリ名を表示。",
            "55555555 (F)": "aa",
            "None (G)": "bbb"
        },
        {
            "No (A)": 2,
            "カテゴリ\n（画面名） (B)": "ログイン",
            "機能名 (C)": "ログイン",
            "機能説明 (D)": "会社名と使用者名を記録する。認証機能は不要。",
            "55555555 (F)": "cccc",
            "None (G)": "ddddd"
        },
        {
            "No (A)": 3,
            "カテゴリ\n（画面名） (B)": "商品選択",
            "機能名 (C)": "画面選択タブ",
            "機能説明 (D)": "「商品選択」か「ライブラリ」を選択できる。",
            "55555555 (F)": "eeee",
            "None (G)": "fffff"
        },
        {
            "No (A)": 4,
            "機能名 (C)": "商品選択ボタン",
            "機能説明 (D)": "果物か野菜かで分類されており、対象商品を押下するとカメラに遷移できる。"
        },
        {
            "No (A)": 6,
            "機能名 (C)": "ログアウト",
            "機能説明 (D)": "ログイン状態をリセットし、スプラッシュに遷移する。ログイン状態で行った作業は保存される。"
        }
        # Bạn có thể thêm các phần tử tiếp theo tương tự
    ]
}

df = pd.DataFrame(data["アプリ機能一覧"])

# Xóa dòng đầu tiên vì nó là header
df = df.drop(0)

# Ghi DataFrame vào file Excel
df.to_excel("data.xlsx", index=False)

# Load workbook
wb = load_workbook("data.xlsx")
ws = wb.active

# Merge cells F và G cho tất cả các dòng
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=6, max_col=7):
    ws.merge_cells(start_row=row[0].row, start_column=6, end_row=row[0].row, end_column=7)

# Lưu file
wb.save("data.xlsx")