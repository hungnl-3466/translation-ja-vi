from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import json
from datetime import datetime

# Hàm chuyển đổi đối tượng datetime thành chuỗi
def datetime_to_string(value):
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    return value
# Đường dẫn tới file Excel
excel_file = "/media/benu/DATA/sun-asterisk/translation-ja-vi/xls_data/Merge_cell_[1356]_ImageProcessingPoC_機能一覧.xlsx"

# Load workbook

# Load workbook
wb = load_workbook(excel_file)

# Khởi tạo một danh sách để lưu trữ dữ liệu
data = []

# Lấy ra tất cả các sheets trong workbook
sheets = wb.sheetnames

# Lặp qua từng sheet để trích xuất dữ liệu
for sheet_name in sheets:
    sheet = wb[sheet_name]
    
    # Khởi tạo một danh sách để lưu trữ dữ liệu từ mỗi dòng trong sheet
    sheet_data = []
    
    # Lặp qua từng dòng trong sheet
    for row in sheet.iter_rows(values_only=True):
        # Khởi tạo một từ điển để lưu trữ dữ liệu từ mỗi ô trong dòng
        row_data = {}
        
        # Lặp qua từng cột trong dòng
        for idx, value in enumerate(row, start=1):
            cell = sheet.cell(row=idx, column=idx)
            if isinstance(value, datetime):
                # Nếu giá trị là datetime, chuyển đổi thành chuỗi
                value = value.strftime("%Y-%m-%d %H:%M:%S")
            
            if value is not None:
                # Kiểm tra xem ô có nằm trong phạm vi merge không
                in_merged_cell = False
                for merged_cell_range in sheet.merged_cells.ranges:
                    min_col, min_row, max_col, max_row = merged_cell_range.min_col, merged_cell_range.min_row, merged_cell_range.max_col, merged_cell_range.max_row
                    if min_col <= cell.column <= max_col and min_row <= cell.row <= max_row:
                        # Lấy giá trị của ô đầu tiên trong phạm vi merge
                        value = merged_cell_range.start_cell.value
                        in_merged_cell = True
                        break
                        
                if not in_merged_cell:
                    column_letter = get_column_letter(cell.column)
                    row_data[f"{sheet.cell(row=1, column=cell.column).value} ({column_letter})"] = value
        
        # Kiểm tra xem dòng có chứa ít nhất một giá trị khác None không trước khi thêm vào danh sách dữ liệu của sheet
        if any(value is not None for value in row_data.values()):
            sheet_data.append(row_data)
    
    # Thêm danh sách dữ liệu của sheet vào danh sách tổng thể nếu có ít nhất một dòng chứa giá trị khác None
    if sheet_data:
        sheet_data_dict = {sheet_name: sheet_data}
        with open('merge_{}.json'.format(sheet_name), 'w', encoding='utf-8') as f:
            json.dump(sheet_data_dict, f,ensure_ascii=False, indent=4)
        data.append({sheet_name: sheet_data})

# In dữ liệu dưới dạng dictionary
print(data)